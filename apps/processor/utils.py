import os
import zipfile
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from django.conf import settings

def process_folder(folder_path):
    processed_dir = os.path.join(settings.MEDIA_ROOT, 'processed')
    summary_dir = os.path.join(settings.MEDIA_ROOT, 'summary')
    os.makedirs(processed_dir, exist_ok=True)
    os.makedirs(summary_dir, exist_ok=True)

    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
    master_filename = f'master_{timestamp}.xlsx'
    summary_filename = f'summary_{timestamp}.xlsx'
    master_path = os.path.join(processed_dir, master_filename)
    summary_path = os.path.join(summary_dir, summary_filename)

   
    master_wb = Workbook()
    default_sheet = master_wb.active  # Keep default sheet for now
    default_sheet.title = "Placeholder"

    summary_wb = Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Sheet Info"

    yellow_fill = PatternFill(start_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)

    
    summary_headers = ["Source File", "Sheet Name","D14","D5"]
    summary_ws.append(summary_headers)
    for cell in summary_ws[1]:
        cell.fill = yellow_fill
        cell.font = bold_font

    real_sheets_added = False  

    for filename in os.listdir(folder_path):
        if filename.endswith('.zip'):
            zip_path = os.path.join(folder_path, filename)
            extract_path = os.path.join(settings.MEDIA_ROOT, 'temp', filename)
            os.makedirs(extract_path, exist_ok=True)

            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_path)

            for root, _, files in os.walk(extract_path):
                for file in files:
                    if file.endswith(('.xlsx', '.xls')):
                        excel_path = os.path.join(root, file)

                        try:
                            xls = pd.ExcelFile(excel_path)
                            for sheet_name in xls.sheet_names:
                                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                                sheet_title = f"{os.path.splitext(file)[0]}_{sheet_name}"[:31]  # Max 31 chars

                                # Add to master workbook
                                ws = master_wb.create_sheet(title=sheet_title)
                                for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                                    for c_idx, value in enumerate(row, start=1):
                                        ws.cell(row=r_idx, column=c_idx, value=value)
                                for c_idx, col_name in enumerate(df.columns, start=1):
                                    cell = ws.cell(row=1, column=c_idx, value=col_name)
                                    cell.font = bold_font

                                # --- Read specific cells and add to summary
                                openpyxl_wb = load_workbook(excel_path, data_only=True)
                                openpyxl_ws = openpyxl_wb[sheet_name]

                                d14 = openpyxl_ws['D14'].value
                                d5 = openpyxl_ws['D5'].value
                                

                                summary_ws.append([file, sheet_name, d14,d5])

                                real_sheets_added = True  

                        except Exception as e:
                            print(f"Error processing {excel_path}: {e}")

    
    if real_sheets_added:
        # If real sheets were added, remove the placeholder sheet
        if "Placeholder" in master_wb.sheetnames:
            std = master_wb["Placeholder"]
            master_wb.remove(std)
    else:
        # If no real sheets were added, write a note
        default_sheet.append(["No valid Excel files found in the uploaded folder."])

    
    master_wb.save(master_path)
    summary_wb.save(summary_path)

    return {
        'master_file': master_path,
        'summary_file': summary_path,
        'original_folder': os.path.basename(folder_path)
    }

